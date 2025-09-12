import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_HEADER_FILL = PatternFill(start_color='44546A', end_color='44546A', fill_type='solid')
BASE_HEADER_FONT = Font(color='FFFFFF', bold=True)

def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def build_posts_report(queryset, filters: dict) -> bytes:
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = 'meta'
    ws_meta.append(['Отчет: Посты'])
    ws_meta.append(['Сгенерировано', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    for k, v in filters.items():
        ws_meta.append([k, ', '.join(v) if isinstance(v, (list, tuple)) else v])

    ws = wb.create_sheet('posts')
    headers = [
        'ID','Тип','Статус','Заголовок','Автор','Создан','Обновлен',
        'Теги','Лайки','Комментарии','Просмотры','Калории','Время_готовки'
    ]
    ws.append(headers)
    for c in ws[1]:
        c.fill = BASE_HEADER_FILL
        c.font = BASE_HEADER_FONT
        c.alignment = Alignment(horizontal='center')

    for p in queryset:
        ws.append([
            p.id,
            p.post_type,
            p.status,
            p.title,
            p.author.username if p.author_id else '',
            p.created_at.strftime('%Y-%m-%d %H:%M'),
            p.updated_at.strftime('%Y-%m-%d %H:%M'),
            ', '.join(p.tags.values_list('name', flat=True)),
            getattr(p, 'likes_count', ''),
            getattr(p, 'comments_count', ''),
            getattr(p, 'views_count', ''),
            p.calories if hasattr(p, 'calories') else '',
            p.cooking_time if hasattr(p, 'cooking_time') else '',
        ])

    autosize(ws_meta)
    autosize(ws)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
